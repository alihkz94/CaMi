#!/usr/bin/env python3
"""
=============================================================================
build_ena_metadata.py — turn the ENA submission workbook into a manifest
-----------------------------------------------------------------------------
PURPOSE
    Read the ENA submission workbook once and write manifests/ena_metadata.tsv,
    the per-sample metadata every downstream analysis stratifies by.

    This is a ONE-TIME script, not a Nextflow process. The workbook is a static
    external file; everything else the STATS step reads is per-sample output
    that changes as the profiling run proceeds. Making this a pipeline task
    would add resume-cache machinery for a file that changes about never. Run
    it by hand when the workbook changes, review the diff like any other
    manifest edit, and commit the result.

    Because it is interactive, it FAILS LOUDLY on a surprise. That is the
    opposite of the rule for pipeline tasks (which degrade gracefully and write
    a SKIPPED report), and it is deliberate: a human is watching this one.

WHERE EACH FIELD ACTUALLY COMES FROM — read before trusting a column
    The workbook is a submission spreadsheet, not a phenotype table. Only two
    of the fields the analysis wants are really in it:

      FROM THE WORKBOOK (authoritative)
        sample_code, ers_accession, disease_status_ena, library_strategy,
        wga_status_ena, sequencing_depth_gb, instrument

      DERIVED FROM THE SAMPLE CODE (the workbook has NO tissue column at all)
        country, location, species, year, individual, tissue, replicate,
        prep suffix, high-coverage flag

      NOT AVAILABLE ANYWHERE YET — written as NA, never guessed
        tumor_purity     would come from the paper's supplementary tables
        btn_lineage      likewise (CedBTN1 / CedBTN2)
        lane_id          the workbook carries no run-level rows at all
        flowcell_id      likewise

    The last group is not an oversight to fix later by inference. An NA here
    means "nobody has measured this for these samples", and a downstream check
    that needs one must say SKIPPED rather than invent a value.

THE JOIN KEY IS sample_code, NOT run_accession
    The workbook's Accessions sheet holds 563 SAMPLE rows (ERS...) and no RUN
    rows (ERR...). There is therefore no ENA run accession anywhere in it, and
    the only key it shares with this pipeline's manifests is the sample code.
    aggregate_counts.py already keys its metadata on sample_code, so they join
    directly.

TISSUE IS CODE-DERIVED, WHICH MAKES THE GRAMMAR LOAD-BEARING
    With no tissue column in the workbook, nothing can cross-check the tissue
    letter. A parser regression here is silent. Three real forms in this cohort
    break a naive "last capital letter is the tissue" rule:

        FRCE17_840F-wga  prep suffix hides the tissue letter
        ENCE17_321H-gp   a SECOND prep suffix, same problem
        PACE17_421H1     trailing replicate digit after the tissue letter
        ICCE19_359F_HC   "_HC" is a coverage annotation; the tissue is the F
        ASCE17_1983      no tissue letter at all — reported, never guessed

    "_HC" is the trap this script was written to close. Read naively the code
    ends in C, and C is not a tissue in the Bruzos codebook at all (it is a
    country letter). All five _HC samples are FOOT, sequenced deeper (60 Gb),
    and belong in a depth subgroup — not in an "unknown tissue" bucket.

INPUT   ENA ScubaCancers Submission.xlsx   (--workbook)
OUTPUT  manifests/ena_metadata.tsv         (--out)

ENV     polars, plus openpyxl (pip install openpyxl==3.1.5)
RUN     python3 scripts/download/build_ena_metadata.py --workbook <xlsx> \
                --out <dataRoot>/manifests/ena_metadata.tsv

IS THIS SCRIPT FOR YOU?
    Only if your reads came from ENA and you hold that study's submission
    workbook. It is one of three ways to give CaMi a design, and the most
    specific. The general way is a samplesheet: see docs/usage.md. Nothing
    in the pipeline requires this file.
VERSION 1.0 (2026-08-21)
=============================================================================
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Final

import polars as pl

# openpyxl reads .xlsx and is needed by THIS SCRIPT ONLY. It is deliberately not
# in env/requirements-stats.txt: that file describes the environment every
# pipeline task runs in, and this script never runs in a task. Putting it there
# would ship an unused dependency inside the stats image and force every cluster
# to re-pull it.
try:
    import openpyxl
except ImportError:
    sys.exit(
        "ERROR: this script needs openpyxl to read the .xlsx workbook.\n"
        "  pip install openpyxl==3.1.5\n"
        "  It is not part of the pipeline environment because no pipeline task "
        "reads a workbook."
    )

# --- the Bruzos biobank codebook -----------------------------------------
# Source: the lab's "BruzosBiobank: Labelling our samples" sheet. A code reads
#   [country:1][location:1][species:2][year:2] _ [number] [tissue:1] [replicate]
# so ENCE16_224H is Spain / Noia / Cerastoderma edule / 2016 / 224 / haemolymph.
COUNTRIES: Final[dict[str, str]] = {
    "A": "Germany", "C": "Croatia", "D": "Denmark", "E": "Spain",
    "F": "France", "H": "Netherlands", "I": "Ireland", "M": "Morocco",
    "N": "Norway", "P": "Portugal", "O": "Poland", "R": "Russia",
    "T": "Italy", "U": "United Kingdom",
}

# (country, location) -> place. The codebook reuses a few letters for the warty
# venus clam work of a different paper (E/C is Carril here, but also appears
# against Cadiz and Vigo there). This cohort is entirely Cerastoderma edule, so
# the cockle meaning is the right one; the collisions are recorded in the
# codebook's own Comments column and do not apply to these 563 samples.
LOCATIONS: Final[dict[tuple[str, str], str]] = {
    ("A", "S"): "Sylt", ("A", "H"): "Harlesiel",
    ("C", "S"): "Split",
    ("D", "N"): "Nykobing Mors", ("D", "V"): "Veno Limfjorden",
    ("E", "A"): "Combarro", ("E", "B"): "Pais Vasco", ("E", "C"): "Carril",
    ("E", "E"): "Espasante", ("E", "G"): "Grove", ("E", "I"): "Camarinas",
    ("E", "L"): "Placeres", ("E", "M"): "Moana", ("E", "N"): "Noia",
    ("E", "O"): "Barallobre", ("E", "P"): "Rio Anllons", ("E", "U"): "Muros",
    ("E", "Y"): "Baiona", ("E", "R"): "Ribeira", ("E", "F"): "Ferrol",
    ("E", "V"): "Marin",
    ("F", "A"): "Arcachon", ("F", "R"): "Roscoff", ("F", "H"): "Le Havre",
    ("F", "O"): "Orne", ("F", "V"): "Baie de Veys", ("F", "B"): "Saint Brieuc",
    ("F", "G"): "Granville", ("F", "C"): "Cabourg",
    ("H", "S"): "Slikken van Viane",
    ("I", "C"): "Cork", ("I", "D"): "Dublin", ("I", "T"): "Inc Beach",
    ("I", "W"): "Westport", ("I", "X"): "Wexford", ("I", "G"): "Carna",
    ("M", "O"): "Oualidia",
    ("N", "B"): "Bodo", ("N", "H"): "Hjeltefjorden",
    ("P", "F"): "Fuzeta", ("P", "L"): "Alvor",
    ("P", "A"): "Ria Formosa (Algarve)", ("P", "V"): "Aveiro",
    ("P", "M"): "Menez Gwen Cage Site",
    ("R", "D"): "Dalnye Zelentsy", ("R", "M"): "Murmansk",
    ("T", "M"): "S. Benedetto",
    ("U", "D"): "Plymouth", ("U", "L"): "Loc Gair", ("U", "S"): "Strollamus",
    ("U", "T"): "Trag Mor", ("U", "G"): "Wales",
}

SPECIES: Final[dict[str, str]] = {
    "CE": "Cerastoderma edule", "CG": "Cerastoderma glaucum",
    "C-": "Cerastoderma spp.", "VV": "Venus verrucosa",
    "PA": "Polititapes aereus", "R-": "Rudititapes spp.",
    "SP": "Scrobicularia plana", "BA": "Bathymodiolus azoricus",
    "ME": "Mytilus edulis",
}

# "F or P" in the codebook: both letters mean foot.
TISSUES: Final[dict[str, str]] = {
    "H": "haemolymph", "A": "adductor muscle", "M": "mantle",
    "D": "digestive/intestine", "G": "gonad", "B": "gill",
    "F": "foot", "P": "foot",
}

# Order matters. Strip the coverage annotation and the prep suffix BEFORE
# looking for the tissue letter, or "_HC" donates a spurious C and "-wga"
# hides the real letter completely.
_HIGH_COVERAGE = re.compile(r"_HC$")
_PREP_SUFFIX = re.compile(r"-([a-z]+)$")
_PREFIX = re.compile(r"^(?P<country>[A-Z])(?P<location>[A-Z])(?P<species>[A-Z-]{2})(?P<year>\d{2})_")
_CODE_GRAMMAR = re.compile(r"^(?P<ind>.*?)(?P<tissue>[A-Z])(?P<rep>\d*)$")

REQUIRED_COLUMNS: Final[tuple[str, ...]] = (
    "Sample code", "Sample type", "Sequencing depth (Gb)",
    "SAMPLE", "INSTRUMENT", "LIBRARY_STRATEGY",
)

# The paper reports NovaSeq 6000; every row of the workbook says HiSeq 2000.
# Recorded and flagged, never silently resolved: the two instruments differ in
# whether index hopping is a real risk (patterned vs non-patterned flowcell),
# which decides whether the index-hopping check is meaningful or not applicable.
PAPER_INSTRUMENT: Final[str] = "Illumina NovaSeq 6000"

# What the cohort must look like if the workbook and the parser are both right.
EXPECTED_DISEASE: Final[dict[str, int]] = {
    "tumor": 61, "matched host": 38, "unmatched host": 2, "healthy": 462,
}
EXPECTED_TISSUE_BY_DISEASE: Final[dict[tuple[str, str], int]] = {
    ("tumor", "H"): 61,
    ("matched host", "F"): 20, ("matched host", "A"): 10, ("matched host", "M"): 8,
    ("unmatched host", "A"): 1, ("unmatched host", "M"): 1,
    ("healthy", "F"): 410, ("healthy", "M"): 17, ("healthy", "H"): 6,
    ("healthy", "none"): 29,
}
EXPECTED_PREP: Final[dict[tuple[str, str], int]] = {
    ("tumor", "native"): 26, ("tumor", "wga"): 35,
    ("matched host", "native"): 26, ("matched host", "wga"): 12,
    ("unmatched host", "native"): 2,
    ("healthy", "native"): 456, ("healthy", "wga"): 6,
}


def split_code(code: str) -> dict:
    """
    Break a sample code into every field it encodes. Never raises: a code that
    does not fit is reported as such, never guessed at.
    """
    rest = code
    is_hc = bool(_HIGH_COVERAGE.search(rest))
    rest = _HIGH_COVERAGE.sub("", rest)

    prep_match = _PREP_SUFFIX.search(rest)
    prep_suffix = prep_match.group(1) if prep_match else ""
    rest = _PREP_SUFFIX.sub("", rest)

    prefix = _PREFIX.match(rest)
    if prefix:
        country = prefix["country"]
        location = prefix["location"]
        species = prefix["species"]
        year = 2000 + int(prefix["year"])
    else:
        country = location = species = ""
        year = None

    grammar = _CODE_GRAMMAR.match(rest)
    if grammar and grammar["tissue"] in TISSUES:
        individual, tissue, replicate = (
            grammar["ind"], grammar["tissue"], grammar["rep"]
        )
    else:
        # No tissue letter (healthy biopsies such as ASCE17_1983), or a
        # trailing capital that is not a known tissue. Both are reported.
        individual, tissue, replicate = rest, "none", ""

    return {
        "individual": individual,
        "tissue_code": tissue,
        "tissue_name": TISSUES.get(tissue, "none"),
        "replicate": replicate,
        "prep_suffix": prep_suffix,
        "is_high_coverage": is_hc,
        "country_code": country,
        "country_name": COUNTRIES.get(country, ""),
        "location_code": f"{country}{location}" if country else "",
        "location_name": LOCATIONS.get((country, location), ""),
        "species_code": species,
        "species_name": SPECIES.get(species, ""),
        "sampling_year": year,
    }


def read_workbook(path: Path) -> list[dict]:
    """Read the Samples sheet, showing what was found before using it."""
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Samples" not in book.sheetnames:
        sys.exit(
            f"ERROR: no 'Samples' sheet in {path}.\n"
            f"       sheets present: {book.sheetnames}"
        )
    rows = list(book["Samples"].iter_rows(values_only=True))
    header = [str(h) if h is not None else "" for h in rows[0]]

    print("Workbook columns found in the 'Samples' sheet:")
    for i, name in enumerate(header):
        print(f"  [{i:2d}] {name}")
    print()

    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        sys.exit(
            "ERROR: the workbook is missing columns this script needs:\n"
            + "".join(f"       {c}\n" for c in missing)
            + "       The workbook layout has changed. Fix the mapping above\n"
            "       rather than working around it downstream."
        )

    at = {name: i for i, name in enumerate(header)}
    return [
        {
            "sample_code": str(r[at["Sample code"]]).strip(),
            "ers_accession": str(r[at["SAMPLE"]]).strip(),
            "disease_status_ena": str(r[at["Sample type"]]).strip(),
            "library_strategy": str(r[at["LIBRARY_STRATEGY"]]).strip(),
            "sequencing_depth_gb": float(r[at["Sequencing depth (Gb)"]]),
            "instrument": str(r[at["INSTRUMENT"]]).strip(),
        }
        for r in rows[1:]
        if r and r[at["Sample code"]]
    ]


def check(label: str, got: dict, want: dict, problems: list[str]) -> None:
    """Compare a tally against what the cohort is known to be."""
    if got == want:
        print(f"  OK   {label}")
        return
    problems.append(label)
    print(f"  FAIL {label}")
    for key in sorted(set(got) | set(want), key=str):
        g, w = got.get(key, 0), want.get(key, 0)
        if g != w:
            print(f"         {key}: workbook says {g}, expected {w}")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Build manifests/ena_metadata.tsv from the ENA submission workbook.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--skip-checks", action="store_true",
                    help="write the table even if the cohort tallies disagree "
                         "(for inspecting a CHANGED workbook, not for normal use)")
    args = ap.parse_args()

    if not args.workbook.exists():
        sys.exit(f"ERROR: workbook not found: {args.workbook}")

    print("=" * 76)
    print(f" workbook : {args.workbook}")
    print(f" out      : {args.out}")
    print("=" * 76)

    records = read_workbook(args.workbook)
    print(f"Rows read: {len(records)}")

    rows = []
    for rec in records:
        parts = split_code(rec["sample_code"])
        strategy = rec["library_strategy"].upper()
        rows.append(
            {
                **rec,
                # WGS = native DNA, WGA = whole-genome amplified (REPLI-g, MDA).
                # The original study excluded WGA from its high-confidence set.
                "wga_status_ena": "wga" if strategy == "WGA" else "native",
                **parts,
                # Recorded per sample so the disagreement travels with the data.
                "instrument_discrepancy": rec["instrument"] != PAPER_INSTRUMENT,
                "instrument_paper": PAPER_INSTRUMENT,
                # Not in the workbook. NA is a statement, not a placeholder.
                "tumor_purity": None,
                "btn_lineage": None,
                "lane_id": None,
                "flowcell_id": None,
            }
        )

    table = pl.DataFrame(rows, strict=False).sort("sample_code")

    # --- does this match the cohort we know? ------------------------------
    print("\nCohort checks (workbook + parser against the known cohort):")
    problems: list[str] = []

    check("total sample count",
          {"n": table.height}, {"n": sum(EXPECTED_DISEASE.values())}, problems)

    check("samples per disease status",
          dict(table.group_by("disease_status_ena").len()
               .select("disease_status_ena", "len").iter_rows()),
          EXPECTED_DISEASE, problems)

    check("tissue per disease status",
          {(d, t): n for d, t, n in
           table.group_by("disease_status_ena", "tissue_code").len()
           .select("disease_status_ena", "tissue_code", "len").iter_rows()},
          EXPECTED_TISSUE_BY_DISEASE, problems)

    check("DNA prep per disease status",
          {(d, p): n for d, p, n in
           table.group_by("disease_status_ena", "wga_status_ena").len()
           .select("disease_status_ena", "wga_status_ena", "len").iter_rows()},
          EXPECTED_PREP, problems)

    hc = table.filter(pl.col("is_high_coverage"))
    check("high-coverage (_HC) samples are foot at 60 Gb",
          {"n": hc.height,
           "all foot": int((hc["tissue_code"] == "F").all()),
           "all 60Gb": int((hc["sequencing_depth_gb"] == 60.0).all())},
          {"n": 5, "all foot": 1, "all 60Gb": 1}, problems)

    if problems and not args.skip_checks:
        print(
            "\nSTOPPING: nothing was written.\n"
            "\n"
            "The tallies above are PRJEB58149's — 563 samples, 61 tumour, 38\n"
            "matched host, 2 unmatched host, 462 healthy. They are checked so a\n"
            "parser regression cannot pass unnoticed, because the workbook has no\n"
            "tissue column and nothing else can cross-check the sample code.\n"
            "\n"
            "  Reproducing PRJEB58149  a mismatch means the workbook changed or\n"
            "                          the parser regressed. Find out which.\n"
            "  A DIFFERENT ENA study   the tallies cannot match. Re-run with\n"
            "                          --skip-checks, and read the report.\n"
        )
        return 1

    # --- report ------------------------------------------------------------
    args.out.parent.mkdir(parents=True, exist_ok=True)
    table.write_csv(args.out, separator="\t", null_value="NA")

    unparsed = table.filter(pl.col("tissue_code") == "none")
    instruments = sorted(set(table["instrument"].to_list()))

    print(f"\nSamples with NO tissue letter in the code: {unparsed.height}")
    print("  These are reported as tissue 'none', never guessed. They can enter")
    print("  a pooled analysis but not a tissue-matched one.")
    for code in unparsed["sample_code"].to_list()[:6]:
        print(f"    {code}")
    if unparsed.height > 6:
        print(f"    ... and {unparsed.height - 6} more")

    print(f"\nINSTRUMENT, as recorded in the workbook: {instruments}")
    if all(i != PAPER_INSTRUMENT for i in instruments):
        print(
            f"  The paper reports {PAPER_INSTRUMENT}. Every workbook row disagrees.\n"
            "  This is UNRESOLVED and it matters for one thing only: index hopping.\n"
            "  A HiSeq 2000 has a non-patterned flowcell and does not hop; a\n"
            "  NovaSeq 6000 has a patterned one and can hop at 0.1-2 %.\n"
            "  Until it is resolved, the index-hopping check treats the workbook\n"
            "  as authoritative and reports itself as not applicable, rather than\n"
            "  running a test whose premise is in doubt."
        )

    print("\nFields that are NOT in this workbook and are written as NA:")
    print("  tumor_purity, btn_lineage   -> the paper's supplementary tables")
    print("  lane_id, flowcell_id        -> ENA run-level metadata (no RUN rows here)")
    print("  tissue                      -> derived from the sample code; the")
    print("                                 workbook has no tissue column at all")

    print(f"\nWROTE -> {args.out}  ({table.height} rows x {table.width} columns)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
